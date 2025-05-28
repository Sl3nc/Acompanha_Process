from pathlib import Path
from time import sleep
from PIL import Image, ImageFont, ImageDraw

import os
import sys
import string
import traceback
from unidecode import unidecode
from abc import abstractmethod, ABCMeta
from collections import OrderedDict

import pandas as pd
from pandas.errors import ParserError
from openpyxl import load_workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from selenium.webdriver.remote.webelement import WebElement
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PySide6.QtWidgets import (
    QMainWindow, QApplication, QWidget, QLabel, QVBoxLayout,QPushButton, QLineEdit
)
from PySide6.QtGui import QPixmap, QIcon, QMovie
from PySide6.QtCore import QThread, QObject, Signal, QSize
from src.window_process import Ui_MainWindow

from tkinter import messagebox
from tkinter.filedialog import askopenfilename
from pathlib import Path

class Arquivo:
    """
    Classe responsável por manipular o arquivo Excel de entrada e saída.
    Permite inserir, validar, ler, alterar e abrir o arquivo de processos.
    """
    NOME_SHEET = 'Deltaprice Judiciais'

    def __init__(self) -> None:
        self.tipos_validos = 'lsx'
        self.caminho = ''
        self.COL_TEXT = 12
        pass

    def inserir(self, button: QPushButton) -> None:
        """
        Abre um diálogo para o usuário selecionar o arquivo de processos.
        Valida o arquivo e atualiza o botão com o nome do arquivo selecionado.
        """
        try:
            self.caminho = askopenfilename()
            if self.caminho == '':
                return
            self.__validar_entrada()
            with open(self.caminho, 'r+'):
                ...
            button.setText(self.caminho[self.caminho.rfind('/') +1:])
            button.setIcon(QPixmap(''))

        except PermissionError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se em aberto em outra janela, favor fecha-la')
        except FileExistsError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
        except Exception as error:
            messagebox.showerror(title='Aviso', message= error)

    def __validar_entrada(self) -> str:
        """
        Valida o caminho do arquivo, verifica o tipo e remove acentos do nome se necessário.
        """
        if self.caminho == '':
            return None
        self.__tipo()
        caminho_uni = unidecode(self.caminho)
        if self.caminho != caminho_uni:
            self.caminho = self.__renomear(caminho_uni)

    def __tipo(self) -> bool:
        """
        Verifica se o arquivo possui a extensão correta.
        """
        if self.caminho[len(self.caminho) -3 :] != self.tipos_validos:
            ultima_barra = self.caminho.rfind('/')
            raise Exception(
                f'Formato inválido do arquivo: {self.caminho[ultima_barra+1:]}')
        return True

    def __renomear(self, caminho) -> str:
        """
        Renomeia o arquivo removendo acentos do nome.
        """
        os.renames(self.caminho, caminho)
        return caminho
    
    def envio_invalido(self) -> bool:
        """
        Retorna True se nenhum arquivo foi selecionado.
        """
        return True if len(self.caminho) == 0 else False

    def ler(self) -> list:
        """
        Lê a coluna E do arquivo Excel e retorna uma lista de números de processos.
        """
        return pd.read_excel(self.caminho, usecols='E').dropna().values.tolist()

    def alterar(self, conteudo: OrderedDict) -> None:
        """
        Altera o conteúdo do arquivo Excel, inserindo os resultados das consultas.
        """
        #TODO Alterar
        wb = load_workbook(self.caminho)
        ws = wb[self.NOME_SHEET]
        for index, lista_movimentos in enumerate(conteudo.values(), 2):
            #print(f'{index} - {lista_movimentos}')
            if lista_movimentos == ['']:
                continue

            if ws.cell(index, self.COL_TEXT).value == None:
                ws.cell(index, self.COL_TEXT, '')

            s = ' **'.join(str(movimento) for movimento in lista_movimentos\
                if movimento[:11] not in str(ws.cell(index, self.COL_TEXT).value))

            ws.cell(index, self.COL_TEXT).value = CellRichText(
                [TextBlock(InlineFont(b=True), s), ws.cell(index, self.COL_TEXT).value]
            )

        wb.save(self.caminho)
          
    def abrir(self) -> None:
        """
        Abre o arquivo Excel gerado para o usuário.
        """
        messagebox.showinfo(title='Aviso', message='Abrindo o arquivo gerado!')
        os.startfile(self.caminho)

class Browser:
    """
    Classe utilitária para criar e configurar uma instância do navegador Chrome via Selenium.
    """
    CHROME_DRIVER_PATH = Path(__file__).parent / 'src'/'drivers'/'chromedriver.exe'

    def make_chrome_browser(self,*options: str, hide = True) -> webdriver.Chrome:
        """
        Cria uma instância do navegador Chrome com as opções fornecidas.
        """
        chrome_options = webdriver.ChromeOptions()

        if options is not None:
            for option in options:
                chrome_options.add_argument(option)

        chrome_service = Service(
            executable_path=str(self.CHROME_DRIVER_PATH),
        )

        browser = webdriver.Chrome(
            service=chrome_service,
            options=chrome_options
        )

        if hide == True:
            browser.set_window_position(-10000,0)

        return browser

class Captcha:
    #TODO CAPTCHA
    WAIT = 2
    NOME_IMG = 'image.png'

    def __init__(self, elem_img: str, elem_input: str, browser: WebElement):
        self.resp = ''
        self.browser = browser
        self.elem_input = elem_input
        self.elem_img = elem_img

    @abstractmethod
    def preencher(self) -> None:
        return NotImplementedError("Implemente este método")

    @abstractmethod
    def imagem(self) -> str:
        return NotImplementedError("Implemente este método")

    def esperar(self) -> None:
        self.resp = ''
        while self.resp == '':
            sleep(self.WAIT)

    def set_valor(self, valor) -> None:
        self.resp = valor

class CaptchaSimples(Captcha):
    def __init__(self, elem_img: str, elem_input: str, browser: WebElement):
        super().__init__(elem_img, elem_input, browser)

    def preencher(self) -> None:
        self.browser.find_element(By.ID, self.elem_input)\
            .send_keys(self.resp)

    def imagem(self) -> str:
        self.browser.find_element(By.CSS_SELECTOR, self.elem_img).screenshot(self.NOME_IMG)
        return self.NOME_IMG

class CaptchaGrid(Captcha):
    def __init__(self, elem_img: str, elem_input: str, browser: WebElement):
        super().__init__(elem_img, elem_input, browser)

        self.cord_resp = [
            [-100,-100], [0,-100], [100,-100],
            [-100,0], [0,0], [100,0],
            [-100,100], [0,100], [100,100]
        ] 

        #Horizontal, Vertical
        self.cord_img = {
            '1':(5, 0),
            '2':(110, 0),
            '3':(220, 0),
            '4':(5, 100),
            '5':(110, 100),
            '6':(220, 100),
            '7':(5, 205),
            '8':(110, 205),
            '9':(220, 205),
        }

    def preencher(self):
        el = self.browser.find_element(By.CSS_SELECTOR, self.CAPTCHA2)
        action = webdriver.common.action_chains.ActionChains(self.browser)

        for var1, var2 in [[0,100], [100,100]]:
            action.move_to_element_with_offset(el, var1, var2)
            action.click()
            action.perform()

    def imagem(self):
        self.browser.find_element(By.CSS_SELECTOR, '#root > div > form > div:nth-child(3) > div > div:nth-child(2) > canvas').screenshot(self.nome_img)

        font = ImageFont.truetype("C:\\Windows\\Fonts\\Verdanab.ttf", 50)
        img = Image.open(self.nome_img)
        draw = ImageDraw.Draw(img)

        for number, posic in self.ref_img.items():
            draw.text(posic, number, 'red', font=font)

        return self.nome_img      

class Tribunal:
    """
    Classe abstrata base para tribunais. Define a interface para consulta de processos.
    """
    TIME_TO_WAIT = 1
    __metaclass__ = ABCMeta

    def __init__(self, browser: WebElement) -> None:
        self.browser = browser
        pass

    @abstractmethod
    def executar(self) -> list[str] | CaptchaSimples:
        """
        Executa a consulta do processo no tribunal.
        """
        raise NotImplementedError("Implemente este método")
    
    @abstractmethod
    def acessar_processo(self, num: str) -> None:
        """
        Acessa a página do processo no tribunal.
        """
        raise NotImplementedError("Implemente este método")
    
    @abstractmethod
    def conteudo(self):
        """
        Extrai o conteúdo relevante do processo.
        """
        raise NotImplementedError("Implemente este método")
    
    def esperar_captcha(self):
        """
        Aguarda o usuário informar o valor do captcha.
        """
        self.valor_captcha = ''
        while self.valor_captcha == '':
            sleep(self.WAIT_CAPTCHA)

    def preencher_captcha(self):
        """
        Preenche o captcha no campo correspondente do site.
        """
        self.browser.find_element(By.ID, self.CAPTCHA)\
            .send_keys(self.valor_captcha)
    
    def set_captcha(self, valor):
        """
        Define o valor do captcha informado pelo usuário.
        """
        self.valor_captcha = valor

class EPROC(Tribunal):
    """
    Implementação do Tribunal EPROC para consulta de processos.
    """
    LINK_BASE = 'https://eproc1g.trf6.jus.br/eproc/externo_controlador.php?acao=processo_consulta_publica'
    INPUT = 'txtNumProcesso'
    CONTULTAR = 'sbmNovo'
    IMG_ELEMENT = '#lblInfraCaptcha > img'
    INPUT_ELEMENT = 'txtInfraCaptcha'
    TABLE_CONTENT = '#divInfraAreaProcesso > table > tbody'
    FRAME_PRINT = [300, 430, 430, 480]

    def __init__(self, browser: WebElement) -> None:
        super().__init__(browser)
        self.CAPTCHA = 'txtInfraCaptcha'
        pass

    def executar(self):
        """
        Executa a consulta no EPROC, tratando captcha se necessário.
        """
        if self.tentar_consulta() == False:
            self.img = self.imagem_captcha()
            return self.img
        os.remove(self.img)
        return self.conteudo()
    
    def acessar_processo(self, num: str) -> None:
        """
        Acessa o processo no EPROC pelo número informado.
        """
        self.browser.get(self.LINK_BASE)
        sleep(1)
        num = num.replace('.','').replace('-','')
        self.browser.find_element(By.ID, self.INPUT).send_keys(num)

    def tentar_consulta(self) -> bool:
        """
        Tenta consultar o processo, verificando se há captcha.
        """
        self.browser.find_element(By.ID, self.CONTULTAR).click()
        sleep(self.TIME_TO_WAIT)
        try:
            #Se dar erro é porque não tem o captcha, senão o contrário
            alert = WebDriverWait(self.browser, self.TIME_TO_WAIT)\
                .until(EC.alert_is_present())
            alert.accept() 
            self.browser.find_element(By.ID, self.CAPTCHA)
        except:
            return True
        return False
        
    def imagem_captcha(self):
        """
        Salva e recorta a imagem do captcha para exibição ao usuário.
        """
        self.browser.save_screenshot(self.NOME_IMG)
        Image.open(self.NOME_IMG).crop([300, 430, 430, 480]).save(self.NOME_IMG)
        return self.NOME_IMG

    def conteudo(self):
        """
        Extrai o conteúdo da tabela de movimentos do processo.
        """
        tbody = self.browser.find_element(By.CSS_SELECTOR, self.TABLE_CONTENT)
        rows = tbody.find_elements(By.TAG_NAME, 'tr')
        rows.pop(0)
        return [x.text[3:] for x in rows if x.text != '']

class PJE(Tribunal):
    """
    Implementação do Tribunal PJE para consulta de processos.
    """
    CLASS_ELEMENTS = 'col-sm-12'
    INPUT = 'fPP:numProcesso-inputNumeroProcessoDecoration:numProcesso-inputNumeroProcesso'
    BTN_PESQUISAR = 'fPP:searchProcessos'
    TABELA_PROCESSO = 'fPP:processosTable:tb'
    TABELA_CONTEUDO = 'j_id134:processoEvento'
    LINK_BASE = 'https://pje-consulta-publica.tjmg.jus.br/'
    LINK_JANELA = 'https://pje-consulta-publica.tjmg.jus.br/pje/ConsultaPublica/DetalheProcessoConsultaPublica/listView.seam?ca'

    def __init__(self, browser) -> None:
        super().__init__(browser)
        pass

    def acessar_processo(self, num_process: str) -> None:
        """
        Acessa o processo no PJE pelo número informado.
        """
        self.browser.get(self.LINK_BASE)
        self.browser.find_element(By.NAME, self.INPUT).send_keys(num_process)

    def executar(self) -> list[str]:
        """
        Executa a consulta no PJE e retorna os movimentos do processo.
        """
        try:
            self.browser.find_element(By.NAME, self.BTN_PESQUISAR).click()
            sleep(self.TIME_TO_WAIT)
            botao_janela = self.browser.find_element(By.ID, self.TABELA_PROCESSO)
            metodo_janela = botao_janela.find_element(By.TAG_NAME, 'a')\
                .get_attribute('onclick')
            link_janela = metodo_janela[metodo_janela.rfind('='):]
            return self.conteudo(link_janela)
        except NoSuchElementException:
            return ['~']

    def conteudo(self, endereco: str) -> list[str]:
        """
        Extrai o conteúdo da tabela de eventos do processo.
        """
        self.browser.get(self.LINK_JANELA + endereco[:len(endereco)-2])
        tbody = self.browser.find_element(By.ID, self.TABELA_CONTEUDO)
        return [x.text for x in tbody.find_elements(By.TAG_NAME, 'span')\
                if x.text != '' and x.text[0].isnumeric()]

class TST(Tribunal):
    LINK_BASE = 'https://consultaprocessual.tst.jus.br/consultaProcessual/consultaTstNumUnica.do?consulta=Consultar&conscsjt=&numeroTst={0}&digitoTst={1}&anoTst={2}&orgaoTst={3}&tribunalTst={4}&varaTst={5}&submit=Consultar'

    TABLE_PROCESSOS = 'body > table > tbody > tr:nth-child(3) > td > table > tbody > tr:nth-child(2) > td > table > tbody > tr:nth-child(32) > td > table > tbody'

    def __init__(self, browser) -> None:
        super().__init__(browser)
        self.cortes_string = [7, 9, 13, 14, 16, 20]
        pass

    def acessar_processo(self, num_process: str) -> None:
        partes = []
        posic_passada = 0
        num_process = num_process.replace('.','').replace('-','')
        for posic in self.cortes_string:
            partes.append(num_process[posic_passada : posic])
            posic_passada = posic

        self.browser.get(self.LINK_BASE.format(
            partes[0],partes[1],partes[2],partes[3],partes[4],partes[5])
        )

    def executar(self):
        try:
            tabela = self.browser.find_element(By.CSS_SELECTOR, self.TABLE_PROCESSOS)
            linhas = tabela.find_elements(By.TAG_NAME, 'tr')
            return self.conteudo(linhas)
        except NoSuchElementException:
            return ['~']
        
    def conteudo(self, rows):
        return [x.text for x in rows[1:] if x.text != '']

class Juiz(QObject):
    """
    Classe responsável por orquestrar a consulta dos processos em diferentes tribunais.
    Utiliza multithreading para não travar a interface gráfica.
    """
    valor = Signal(str)
    progress = Signal(int)
    fim = Signal(OrderedDict)
    WAIT_CAPTCHA = 2

    def __init__(self, num_process: list[str]) -> None:
        super().__init__()
        self.num_process = num_process
        self.browser = Browser().make_chrome_browser(hide=False)
        self.ref = {
            '13': PJE(self.browser),
            '01': EPROC(self.browser),
            '03': TST(self.browser)
        }

    def pesquisar(self):
        """
        Realiza a pesquisa dos processos, emitindo sinais de progresso e resultado.
        """
        try:
            ref = OrderedDict(
                [(str(x[0])[:25], '') for x in self.num_process]
            )
            for index, num in enumerate(self.num_process, 1):
                num = str(num[0])[:25]
                if num == None:
                    num = ''
                self.processo(ref, num)
                self.progress.emit(index)

        except NoSuchElementException as err:
            traceback.print_exc()
            messagebox.showerror('Aviso', f'Erro na busca por elementos no site, favor comunicar o desenvolvedor. Erro: \n {err}')
        except Exception as err:
            traceback.print_exc()
            messagebox.showerror('Aviso', err)
        finally:
            self.browser.close()
            self.fim.emit(ref)

    def processo(self, ref, num):
        """
        Realiza a consulta de um processo específico, tratando captcha se necessário.
        """
        tribunal_atual = self.__apurar(num)
        if tribunal_atual == None:
            ref[num] = ['']
        else:
            #TODO PESQUISA PROCESSO
            tribunal_atual.acessar_processo(num)
            resp = tribunal_atual.executar()
            while type(resp) == CaptchaSimples:
                self.captcha = resp
                self.valor.emit(self.captcha.imagem())
                self.captcha.esperar()
                self.captcha.preencher()
                resp = tribunal_atual.executar()
            ref[num] = resp
    
    def __apurar(self, num:str) -> Tribunal:
        """
        Determina o tribunal responsável pelo processo a partir do número.
        """
        if len(num) < 16:
            return None 
        for key, value in self.ref.items():
            if key == num.replace('-','').replace('.','')[14:16]:
                return value 
        return None

    def set_captcha(self, valor) -> None:
        """
        Define o valor do captcha para o tribunal atual.
        """
        self.tribunal_atual.set_captcha(valor)

class MainWindow(QMainWindow, Ui_MainWindow):
    """
    Classe principal da interface gráfica. Gerencia as interações do usuário e o fluxo do programa.
    """
    MAX_PROGRESS = 100

    def __init__(self, parent = None):
        super().__init__(parent)
        self.setupUi(self)
        self.text_aviso = [
            'Os tribunais dos seguintes processos ainda não foram implementados no programa:',
            'Os números a seguir não foram encontrados em seus respectivos sites:'
        ]

        self.file = Arquivo()
        self.setWindowTitle('Consulta Processual')
        self.setWindowIcon((QIcon(
            (Path(__file__).parent/'src'/'imgs'/'procss-icon.ico').__str__())))
        self.logo.setPixmap(QPixmap(
            (Path(__file__).parent/'src'/'imgs'/'procss-hori.png').__str__()))
        icon = QIcon()
        icon.addFile(
            (Path(__file__).parent/'src'/'imgs'/'upload-icon.png').__str__(),
            QSize(),
            QIcon.Mode.Normal,
            QIcon.State.Off
        )
        self.pushButton_2.setIcon(icon)
        self.movie = QMovie(
            (Path(__file__).parent/'src'/'imgs'/'load.gif').__str__()
        )
        self.gif_load.setMovie(self.movie)

        self.pushButton_2.clicked.connect(
            lambda: self.file.inserir(self.pushButton_2)
        )
        self.pushButton.clicked.connect(self.hard_work)
        self.enviar_captcha.clicked.connect(self.enviar_resp)

    def hard_work(self):
        """
        Inicia o processamento dos processos em uma thread separada.
        """
        try:
            if self.file.envio_invalido():
                raise Exception('Favor anexar seu relatório de processos')
            
            list_processos = self.file.ler()
            self.posicao = self.MAX_PROGRESS / len(list_processos)

            self.exec_load(True)
            self.pushButton.setDisabled(True)

            self.juiz = Juiz(list_processos)
            self._thread = QThread()

            self.juiz.moveToThread(self._thread)
            self._thread.started.connect(self.juiz.pesquisar)
            self.juiz.fim.connect(self._thread.quit)
            self.juiz.fim.connect(self._thread.deleteLater)
            self.juiz.fim.connect(self.encerramento)
            self._thread.finished.connect(self.juiz.deleteLater)
            self.juiz.valor.connect(self.to_captcha) 
            self.juiz.progress.connect(self.to_progress)

            self._thread.start()  

        except ParserError:
            messagebox.showerror(title='Aviso', message= 'Erro ao ler o arquivo, certifique-se de ter inserido o arquivo correto')
        except Exception as err:
            traceback.print_exc()
            messagebox.showerror('Aviso', err)

    def encerramento(self, result: OrderedDict):
        """
        Finaliza o processamento, exibe avisos e abre o arquivo gerado.
        """
        #TODO encerramento
        invalidos = self.filtro(result)
        for index, i in enumerate(invalidos):
            if len(i) != 0:
                messagebox.showwarning('Aviso', \
                    f'{self.text_aviso[index]} \n {'\n'.join(f'- {x}' for x in i)}')
            
        self.file.alterar(result)
        self.file.abrir()

        self.exec_load(False, 0)
        self.pushButton.setDisabled(False)

    def filtro(self, result: OrderedDict):
        """
        Filtra processos inválidos ou não encontrados.
        """
        falhas = []
        invalido = []
        for key, value in result.items():
            if value == ['']:
                falhas.append(key)
            elif value == ['~']:
                invalido.append(key)
                result[key] = ''
        return [falhas , invalido]
    
    def to_captcha(self, nome_img):
        """
        Exibe a imagem do captcha para o usuário.
        """
        self.label_5.setPixmap(QPixmap(nome_img))
        self.exec_load(False, 2)

    def to_progress(self, valor):
        """
        Atualiza a barra de progresso.
        """
        self.progressBar.setValue(self.posicao * valor)

    def enviar_resp(self):
        """
        Envia a resposta do captcha informada pelo usuário.
        """
        self.juiz.set_captcha(self.lineEdit.text())
        self.lineEdit.setText('')
        self.exec_load(True)

    def exec_load(self, action: bool, to = 1):
        """
        Controla a exibição do GIF de carregamento e troca de telas.
        """
        if action == True:
            self.movie.start()
            self.stackedWidget.setCurrentIndex(to)
        else:
            self.movie.stop()
            self.stackedWidget.setCurrentIndex(to)


if __name__ == '__main__':
    """
    Ponto de entrada do programa. Inicializa a aplicação Qt.
    """
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec()