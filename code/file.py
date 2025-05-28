from openpyxl.cell.rich_text import TextBlock, CellRichText
from tkinter.filedialog import askopenfilename
from openpyxl.cell.text import InlineFont
from PySide6.QtWidgets import QPushButton
from collections import OrderedDict
from openpyxl import load_workbook
from PySide6.QtGui import QPixmap
from os import renames, startfile
from unidecode import unidecode
from tkinter import messagebox
import pandas as pd

class Arquivo:
    """
    Classe responsável por manipular o arquivo Excel de entrada e saída.
    Permite inserir, validar, ler, alterar e abrir o arquivo de processos.
    """
    NOME_SHEET = 'Deltaprice Judiciais'

    def __init__(self) -> None:
        self.tipos_validos = 'lsx'
        self.caminho = ''
        self.COL_TEXT = 12
        pass

    def inserir(self, button: QPushButton) -> None:
        """
        Abre um diálogo para o usuário selecionar o arquivo de processos.
        Valida o arquivo e atualiza o botão com o nome do arquivo selecionado.
        """
        try:
            self.caminho = askopenfilename()
            if self.caminho == '':
                return
            self.__validar_entrada()
            with open(self.caminho, 'r+'):
                ...
            button.setText(self.caminho[self.caminho.rfind('/') +1:])
            button.setIcon(QPixmap(''))

        except PermissionError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se em aberto em outra janela, favor fecha-la')
        except FileExistsError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
        except Exception as error:
            messagebox.showerror(title='Aviso', message= error)

    def __validar_entrada(self) -> str:
        """
        Valida o caminho do arquivo, verifica o tipo e remove acentos do nome se necessário.
        """
        if self.caminho == '':
            return None
        self.__tipo()
        caminho_uni = unidecode(self.caminho)
        if self.caminho != caminho_uni:
            self.caminho = self.__renomear(caminho_uni)

    def __tipo(self) -> bool:
        """
        Verifica se o arquivo possui a extensão correta.
        """
        if self.caminho[len(self.caminho) -3 :] != self.tipos_validos:
            ultima_barra = self.caminho.rfind('/')
            raise Exception(
                f'Formato inválido do arquivo: {self.caminho[ultima_barra+1:]}')
        return True

    def __renomear(self, caminho) -> str:
        """
        Renomeia o arquivo removendo acentos do nome.
        """
        renames(self.caminho, caminho)
        return caminho
    
    def envio_invalido(self) -> bool:
        """
        Retorna True se nenhum arquivo foi selecionado.
        """
        return True if len(self.caminho) == 0 else False

    def ler(self) -> list:
        """
        Lê a coluna E do arquivo Excel e retorna uma lista de números de processos.
        """
        return pd.read_excel(self.caminho, usecols='E').dropna().values.tolist()

    def alterar(self, conteudo: OrderedDict) -> None:
        """
        Altera o conteúdo do arquivo Excel, inserindo os resultados das consultas.
        """
        wb = load_workbook(self.caminho)
        ws = wb[self.NOME_SHEET]
        for index, lista_movimentos in enumerate(conteudo.values(), 2):
            #print(f'{index} - {lista_movimentos}')
            if lista_movimentos == ['']:
                continue

            if ws.cell(index, self.COL_TEXT).value == None:
                ws.cell(index, self.COL_TEXT, '')

            s = ' **'.join(str(movimento) for movimento in lista_movimentos\
                if movimento[:11] not in str(ws.cell(index, self.COL_TEXT).value))

            ws.cell(index, self.COL_TEXT).value = CellRichText(
                [TextBlock(InlineFont(b=True), s), ws.cell(index, self.COL_TEXT).value]
            )

        wb.save(self.caminho)
          
    def abrir(self) -> None:
        """
        Abre o arquivo Excel gerado para o usuário.
        """
        messagebox.showinfo(title='Aviso', message='Abrindo o arquivo gerado!')
        startfile(self.caminho)