from PySide6.QtWidgets import QMainWindow, QApplication, QPushButton
from PySide6.QtGui import QPixmap, QIcon
from PySide6.QtCore import QSize
from src.window_process import Ui_MainWindow
from pathlib import Path
from time import sleep
from tkinter import messagebox
from tkinter.filedialog import askopenfilename

import traceback
import sys
import os
import string
from unidecode import unidecode
import keyboard
from abc import abstractmethod, ABCMeta

import pandas as pd
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

def resource_path(relative_path):
        base_path = getattr(
            sys,
            '_MEIPASS',
            os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(base_path, relative_path)

class Arquivo:
    def __init__(self) -> None:
        self.tipos_validos = 'lsx'
        self.caminho = ''
        self.COL_TEXT = 2
        self.COL_NUM = 1
        pass

    def inserir(self, button: QPushButton):
        try:
            caminho = askopenfilename()
            if caminho == '':
                return None
            self.caminho = self.__validar_entrada(caminho)
            button.setText(caminho[caminho.rfind('/') +1:])
            button.setIcon(QPixmap(''))

        except PermissionError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado apresenta-se em aberto em outra janela, favor fecha-la')
        except FileExistsError:
            messagebox.showerror(title='Aviso', message= 'O arquivo selecionado já apresenta uma versão sem acento, favor usar tal versão ou apagar uma delas')
        except Exception as error:
            messagebox.showerror(title='Aviso', message= error)

    def __validar_entrada(self, caminho):
        if any(c not in string.ascii_letters for c in caminho):
            caminho = self.__formato_ascii(caminho)

        self.__tipo(caminho)
        return caminho

    def __tipo(self, caminho):
        if caminho[len(caminho) -3 :] != self.tipos_validos:
            ultima_barra = caminho.rfind('/')
            raise Exception(
                f'Formato inválido do arquivo: {caminho[ultima_barra+1:]}')

    def __formato_ascii(self, caminho) -> str:
        caminho_uni = unidecode(caminho)
        os.renames(caminho, caminho_uni)
        return caminho_uni
    
    def envio_invalido(self) -> bool:
        return True if len(self.caminho) == 0 else False

    def ler(self) -> list:
        return pd.read_excel(self.caminho, usecols='A', header=None)[0]\
            .values.tolist()

    def alterar(self, conteudo: dict) -> None:
        wb = load_workbook(self.caminho)
        ws = wb.active
        for index, lista_movimentos in enumerate(conteudo.values(), 1):
            valor_novo = ''
            for movimento in lista_movimentos:
                if movimento not in ws.cell(index, self.COL_NUM).value:
                    valor_novo = f'{valor_novo} §#§ {movimento}'

            if ws.cell(index, self.COL_TEXT).value == None:
                ws.cell(index, self.COL_TEXT, '')
            ws.cell(index, self.COL_TEXT).value = \
                ws.cell(index, self.COL_TEXT).value + valor_novo

        wb.save(self.caminho)

    def abrir(self) -> None:
        messagebox.showinfo(title='Aviso', message='Abrindo o arquivo gerado!')
        os.startfile(self.caminho)

# Chrome Options
# https://peter.sh/experiments/chromium-command-line-switches/

class Browser:
    ROOT_FOLDER = Path(__file__).parent
    CHROME_DRIVER_PATH = ROOT_FOLDER / 'src' / 'drivers' / 'chromedriver.exe'

    def __init__(self) -> None:
        pass

    def make_chrome_browser(self,*options: str) -> webdriver.Chrome:
        chrome_options = webdriver.ChromeOptions()

        # chrome_options.add_argument('--headless')
        if options is not None:
            for option in options:
                chrome_options.add_argument(option)

        chrome_service = Service(
            executable_path=str(self.CHROME_DRIVER_PATH),
        )

        browser = webdriver.Chrome(
            service=chrome_service,
            options=chrome_options
        )

        return browser

class Tribunal:
    TIME_TO_WAIT = 1
    __metaclass__ = ABCMeta

    def __init__(self, options = ()) -> None:
        self.browser = Browser().make_chrome_browser(*options)
        pass

    @abstractmethod
    def exec(self, num_processo):
        raise NotImplementedError("Implemente este método")
    
class EPROC(Tribunal):
    LINK_BASE = 'https://eproc1g.trf6.jus.br/eproc/externo_controlador.php?acao=processo_consulta_publica'
    INPUT = 'txtNumProcesso'
    CAPTCHA = 'txtInfraCaptcha'
    CONTULTAR = 'sbmNovo'
    TABLE_CONTENT = '#divInfraAreaProcesso > table > tbody'
    WAIT_CAPTCHA = 2
    FRAME_PRINT = [300, 430, 430, 480]
    NOME_PRINT = 'image.png'
    #1080458-33.2021.4.01.3800

    def __init__(self) -> None:
        super().__init__()
        pass

    def exec(self, num_processo):
        self.browser.get(self.LINK_BASE)
        self.browser.find_element(By.ID, self.INPUT)\
            .send_keys(num_processo.replace('-','').replace('.',''))
        self.tentar_consulta()
        return [
            conteudo.text[3:] for conteudo \
                in self.__valor_janela() if conteudo.text != ''
        ]
       
    def __valor_janela(self):
        tbody = self.browser.find_element(By.CSS_SELECTOR, self.TABLE_CONTENT)
        rows = tbody.find_elements(By.TAG_NAME, 'tr')
        rows.pop(0)
        return rows

    def tentar_consulta(self):
        self.browser.find_element(By.ID, self.CONTULTAR).click()
        keyboard.press_and_release('enter')
        sleep(self.TIME_TO_WAIT)
        try:
            #se o captcha aparece, esperar para preenchê-lo
            if self.browser.find_element(By.ID, self.CAPTCHA).is_displayed():
                return self.preencher_captcha()
        except:
            return None

    def preencher_captcha(self):
        while len(self.browser.find_element(By.ID, self.CAPTCHA).get_attribute('value')) != 4:
            sleep(self.WAIT_CAPTCHA)
        self.tentar_consulta()

class PJE(Tribunal):
    CLASS_ELEMENTS = 'col-sm-12'
    INPUT = 'fPP:numProcesso-inputNumeroProcessoDecoration:numProcesso-inputNumeroProcesso'
    BTN_PESQUISAR = 'fPP:searchProcessos'
    JANELA_PROCESSO = '#fPP\\:processosTable\\:632256959\\:j_id245 > a'
    TABELA_CONTEUDO = 'j_id134:processoEvento:tb'
    LINK_BASE = 'https://pje-consulta-publica.tjmg.jus.br/'
    LINK_JANELA = 'https://pje-consulta-publica.tjmg.jus.br/pje/ConsultaPublica/DetalheProcessoConsultaPublica/listView.seam?ca'
    #5147698-10.2023.8.13.0024

    def __init__(self) -> None:
        super().__init__(('--headless'))
        pass

    def exec(self, num_processo) -> str:
        self.browser.get(self.LINK_BASE)

        self.browser.find_element(By.NAME, self.INPUT)\
            .send_keys(num_processo)


        self.browser.find_element(By.NAME, self.BTN_PESQUISAR).click()

        sleep(self.TIME_TO_WAIT)

        metodo_janela = self.browser.find_element(By.CSS_SELECTOR, self.JANELA_PROCESSO).get_attribute('onclick')

        link_janela = metodo_janela[metodo_janela.rfind('='):]

        return [
            conteudo.text for conteudo \
                in self.__valor_janela(link_janela) if conteudo.text != ''
        ]

    def __valor_janela(self, endereco: str):
        self.browser.get(self.LINK_JANELA + endereco[:len(endereco)-2])

        sleep(self.TIME_TO_WAIT)

        tbody = self.browser.find_element(By.ID, self.TABELA_CONTEUDO)
        return tbody.find_elements(By.TAG_NAME, 'span')

class Juiz:
    def __init__(self) -> None:
        self.ref = {
            '13': PJE(),
            '01': EPROC()
        }
        pass

    def pesquisar(self, num_processos: list) -> dict:
        ref = {}
        for num in num_processos:
            tribunal = self.__apurar(str(num))
            if tribunal == None:
                ref[num] = None
            else:
                ref[num] = tribunal.exec(str(num))
        return ref

    def __apurar(self, num:str) -> Tribunal:
        for key, value in self.ref.items():
            if key == num.replace('-','').replace('.','')[14:16]:
                return value 
        return None
    
class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent = None):
        super().__init__(parent)
        self.setupUi(self)
        self.file = Arquivo()

        self.logo.setPixmap(QPixmap(resource_path("src\\imgs\\procss-icon.ico")))
        icon = QIcon()
        icon.addFile(resource_path("src\\imgs\\upload-icon.png"), QSize(), QIcon.Mode.Normal, QIcon.State.Off)
        self.pushButton_2.setIcon(icon)

        self.pushButton_2.clicked.connect(
            lambda: self.file.inserir(self.pushButton_2)
        )

        self.pushButton.clicked.connect(
            lambda: self.buscar()
        )

    def buscar(self):
        try:
            if self.file.envio_invalido():
                raise Exception('Favor anexar seu relatório de processos')

            juiz = Juiz()        
            num_processos = self.file.ler()
            result = juiz.pesquisar(num_processos)
            invalidos = self.filtra_invalido(result)

            if len(invalidos) != 0:
                for key in invalidos:
                    result.pop(key)
                    
                messagebox.showerror('Aviso', \
                    f'Os tribunais dos seguintes processos ainda não foram implementados no programa: \n\n \
                        {'\n - '.join(str(x) for x in invalidos)}')

            self.file.alterar(result)
            self.file.abrir()
        except Exception as err:
            traceback.print_exc()
            messagebox.showerror('Aviso', err)

    def filtra_invalido(self, result: dict):
        falhas = []
        for key, value in result.items():
            if value == None:
                falhas.append(key)
        return falhas

    def loading(self):
        ...

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    app.exec()